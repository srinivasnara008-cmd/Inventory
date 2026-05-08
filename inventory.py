import requests
from requests.auth import HTTPBasicAuth
import json
import urllib3
import subprocess
import os
import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from pathlib import Path
import time

# Suppress InsecureRequestWarning
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Configuration
MAX_RETRIES = 3
TIMEOUT = 60
RETRY_DELAY = 2

def ping_ip(ip_address):
    """Ping an IP address to verify connectivity"""
    try:
        response = subprocess.run(
            ["ping", "-n", "4", ip_address],  # Use "-c 4" for Linux/macOS
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            timeout=10
        )
        
        if response.returncode == 0:
            print(f"✓ Ping to {ip_address} successful")
            return True
        else:
            print(f"✗ Ping to {ip_address} failed")
            return False
    except subprocess.TimeoutExpired:
        print(f"✗ Ping to {ip_address} timed out")
        return False
    except Exception as e:
        print(f"✗ Error pinging {ip_address}: {e}")
        return False


def fetch_input():
    """Fetch ILO credentials from Excel file with better path handling"""
    try:
        # Try multiple possible paths
        possible_paths = [
            "System_IPs.xlsx",  # Current directory
            Path.home() / "Desktop" / "Inventory" / "System_IPs.xlsx",  # User's Desktop
            Path(__file__).parent / "System_IPs.xlsx",  # Same directory as script
        ]
        
        file_path = None
        for path in possible_paths:
            if os.path.exists(path):
                file_path = path
                break
        
        if not file_path:
            raise FileNotFoundError(f"System_IPs.xlsx not found in any of the expected locations")
        
        print(f"Loading IPs from: {file_path}")
        df = pd.read_excel(file_path)
        
        # Validate required columns
        required_cols = ['ILO_IP', 'ILO_USERNAME', 'ILO_PASSWORD']
        for col in required_cols:
            if col not in df.columns:
                raise ValueError(f"Missing required column: {col}")
        
        result_dict = df.set_index('ILO_IP')[['ILO_USERNAME', 'ILO_PASSWORD']].to_dict(orient='index')
        return result_dict
    except Exception as e:
        print(f"✗ Error fetching input: {e}")
        sys.exit(1)


def risget(url, ilo_user, ilo_password, retry_count=0):
    """Make REST API GET request with retry logic"""
    try:
        response = requests.get(
            url,
            auth=HTTPBasicAuth(str(ilo_user), str(ilo_password)),
            verify=False,
            timeout=TIMEOUT
        )
        print(f"Response status: {response.status_code}")
        return response
    except requests.exceptions.Timeout:
        print(f"✗ Timeout requesting {url}")
        if retry_count < MAX_RETRIES:
            print(f"  Retrying ({retry_count + 1}/{MAX_RETRIES})...")
            time.sleep(RETRY_DELAY)
            return risget(url, ilo_user, ilo_password, retry_count + 1)
        return None
    except requests.exceptions.ConnectionError:
        print(f"✗ Connection error for {url}")
        if retry_count < MAX_RETRIES:
            print(f"  Retrying ({retry_count + 1}/{MAX_RETRIES})...")
            time.sleep(RETRY_DELAY)
            return risget(url, ilo_user, ilo_password, retry_count + 1)
        return None
    except Exception as e:
        print(f"✗ Error in API request: {e}")
        if retry_count < MAX_RETRIES:
            print(f"  Retrying ({retry_count + 1}/{MAX_RETRIES})...")
            time.sleep(RETRY_DELAY)
            return risget(url, ilo_user, ilo_password, retry_count + 1)
        return None


def success_response(response):
    """Validate if response is successful"""
    if response is None:
        print("✗ Response is None")
        return False
    if response.status_code in [200, 201]:
        return True
    print(f"✗ Invalid response status: {response.status_code}")
    return False


def get_platform(response):
    """Extract platform/model information"""
    if success_response(response):
        try:
            obj_dict = json.loads(response.text)
            platform = obj_dict.get("Model", "NA")
            return platform
        except json.JSONDecodeError:
            print("✗ Failed to parse JSON response")
            return "NA"
    return "NA"


def get_storage_id_count(response):
    """Get list of storage controllers"""
    if success_response(response):
        try:
            obj_dict = json.loads(response.text)
            count = obj_dict.get("Members@odata.count", 0)
            if count > 0:
                members = obj_dict.get("Members", [])
                return members if members else None
        except json.JSONDecodeError:
            print("✗ Failed to parse storage JSON")
    return None


def get_controller_name(response, file):
    """Extract controller name and firmware info"""
    if success_response(response):
        try:
            obj_dict = json.loads(response.text)
            controller_name = obj_dict.get("Name", "No Controller detected")
            firmware = obj_dict.get("FirmwareVersion", "NA")
            controller_model = obj_dict.get("Model", "NA")
            
            file.write("-------------------------------------\n")
            file.write(f"Controller Name: {controller_name}\n")
            file.write(f"Controller Model: {controller_model}\n")
            file.write(f"Firmware Version: {firmware}\n")
            file.write("-------------------------------------\n")
            
            return {
                'name': controller_name,
                'model': controller_model,
                'firmware': firmware
            }
        except json.JSONDecodeError:
            print("✗ Failed to parse controller JSON")
    return {'name': 'NA', 'model': 'NA', 'firmware': 'NA'}


def get_drives_list(response):
    """Get list of drives from storage controller"""
    if success_response(response):
        try:
            obj_dict = json.loads(response.text)
            drives_list = obj_dict.get("Drives", [])
            if not drives_list:
                return None
            
            drive_list = []
            for drive in drives_list:
                if "@odata.id" in drive:
                    drive_list.append(drive["@odata.id"])
            return drive_list if drive_list else None
        except json.JSONDecodeError:
            print("✗ Failed to parse drives JSON")
    return None


def fetch_drive_details(response, file):
    """Extract drive details including new fields"""
    if success_response(response):
        try:
            obj_dict = json.loads(response.text)
            drive_model = obj_dict.get("Model", "NA")
            drive_sn = obj_dict.get("SerialNumber", "NA")
            drive_capacity = obj_dict.get("CapacityBytes", "NA")
            drive_status = obj_dict.get("Status", {}).get("State", "NA")
            drive_firmware = obj_dict.get("Revision", "NA")
            drive_type = obj_dict.get("MediaType", "NA")
            
            # Convert capacity from bytes to GB
            if drive_capacity != "NA":
                try:
                    capacity_gb = int(drive_capacity) / (1024**3)
                    drive_capacity = f"{capacity_gb:.2f} GB"
                except:
                    drive_capacity = "NA"
            
            drive_data = {
                'model': drive_model,
                'serial': drive_sn,
                'capacity': drive_capacity,
                'status': drive_status,
                'firmware': drive_firmware,
                'type': drive_type
            }
            
            file.write("------------------------------\n")
            file.write(f"Model: {drive_model} | Serial: {drive_sn}\n")
            file.write(f"Type: {drive_type} | Capacity: {drive_capacity}\n")
            file.write(f"Status: {drive_status} | Firmware: {drive_firmware}\n")
            
            return drive_data
        except json.JSONDecodeError:
            print("✗ Failed to parse drive JSON")
    
    return {
        'model': 'NA',
        'serial': 'NA',
        'capacity': 'NA',
        'status': 'NA',
        'firmware': 'NA',
        'type': 'NA'
    }


def start_inventory(ilo_ip, ilo_user, ilo_password):
    """Main inventory collection function with robust error handling"""
    try:
        file = open("Report.txt", "a")
        drives_list = []
        
        https = "https://"
        base = "/redfish/v1/Systems/1"
        url = https + ilo_ip + base
        
        # Fetch platform info
        response = risget(url, ilo_user, ilo_password)
        if response is None:
            print(f"✗ Failed to fetch system info for {ilo_ip}")
            file.write(f"\n✗ Failed to fetch data from {ilo_ip}\n")
            file.close()
            return None
        
        platform = get_platform(response)
        
        file.write("\n" + "="*100 + "\n")
        file.write(f"Data from server: ILO IP: {ilo_ip} | Platform: {platform}\n")
        file.write("="*100 + "\n")
        
        # Fetch storage info
        response = risget(url + "/Storage", ilo_user, ilo_password)
        storage_id_list = get_storage_id_count(response)
        
        if not storage_id_list:
            print(f"✗ No storage devices found or unable to fetch for {ilo_ip}")
            file.write("No storage information available\n")
            file.close()
            return None
        
        # Process each storage controller
        for sid in storage_id_list:
            controller_url = https + ilo_ip + sid["@odata.id"]
            response = risget(controller_url, ilo_user, ilo_password)
            
            if response is None:
                print(f"✗ Failed to fetch controller info from {controller_url}")
                continue
            
            controller_info = get_controller_name(response, file)
            file.write("Drive Details:\n")
            file.write("-" * 60 + "\n")
            
            # Fetch drives
            drives = get_drives_list(response)
            
            if drives:
                print(f"Found {len(drives)} drives")
                for drive_url in drives:
                    full_drive_url = https + ilo_ip + drive_url
                    drive_response = risget(full_drive_url, ilo_user, ilo_password)
                    
                    if drive_response:
                        drive_data = fetch_drive_details(drive_response, file)
                        drives_list.append({
                            'ilo_ip': ilo_ip,
                            'platform': platform,
                            'controller_name': controller_info['name'],
                            'controller_model': controller_info['model'],
                            'controller_firmware': controller_info['firmware'],
                            **drive_data
                        })
            else:
                print(f"No drives found for controller")
                file.write("No drives detected\n")
        
        file.write("\n" + "="*100 + "\n")
        file.close()
        
        # Save to Excel
        if drives_list:
            save_to_inventory_excel(drives_list)
            return drives_list
        else:
            print(f"✗ No inventory data collected for {ilo_ip}")
            return None
            
    except Exception as e:
        print(f"✗ Error in start_inventory: {e}")
        try:
            file.close()
        except:
            pass
        return None


def save_to_inventory_excel(drives_list, filename="Inventory.xlsx"):
    """Save inventory data to Excel with enhanced columns"""
    try:
        # Create DataFrame with all columns
        columns = [
            "ILO_IP", "Platform", "Controller_Name", "Controller_Model",
            "Controller_Firmware", "Drive_Model", "Drive_Serial",
            "Drive_Capacity", "Drive_Status", "Drive_Firmware", "Drive_Type"
        ]
        
        rows = []
        for drive in drives_list:
            rows.append([
                drive.get('ilo_ip', 'NA'),
                drive.get('platform', 'NA'),
                drive.get('controller_name', 'NA'),
                drive.get('controller_model', 'NA'),
                drive.get('controller_firmware', 'NA'),
                drive.get('model', 'NA'),
                drive.get('serial', 'NA'),
                drive.get('capacity', 'NA'),
                drive.get('status', 'NA'),
                drive.get('firmware', 'NA'),
                drive.get('type', 'NA')
            ])
        
        df = pd.DataFrame(rows, columns=columns)
        
        if not os.path.exists(filename):
            df.to_excel(filename, index=False)
            print(f"✓ Created new inventory file: {filename}")
        else:
            with pd.ExcelWriter(filename, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                df.to_excel(writer, index=False, header=False, startrow=writer.sheets["Sheet1"].max_row)
            print(f"✓ Updated inventory file: {filename}")
        
        # Format Excel file
        wb = load_workbook(filename)
        ws = wb.active
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = Font(bold=True)
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        # Format header
        for col in ws[1]:
            col.font = header_font
            col.alignment = align_center
            col.border = thin_border
        
        # Format data cells
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(columns)):
            for cell in row:
                cell.alignment = align_center
                cell.border = thin_border
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
        
        wb.save(filename)
        print(f"✓ Formatted inventory file successfully")
        
    except Exception as e:
        print(f"✗ Error saving to Excel: {e}")

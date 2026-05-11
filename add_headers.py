"""
Script to create or update System_IPs.xlsx with proper headers and formatting.
This ensures the inventory.py script can read credentials correctly.
"""

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

def create_excel_template(filename="System_IPs.xlsx"):
    """Create a new System_IPs.xlsx template with headers and sample data"""
    try:
        # Define column headers
        headers = ["ILO_IP", "ILO_USERNAME", "ILO_PASSWORD"]
        
        # Sample data (user should replace these with actual values)
        sample_data = [
            ["192.168.1.100", "admin", "password123"],
            ["192.168.1.101", "admin", "password123"],
            ["192.168.1.102", "admin", "password123"],
        ]
        
        # Create DataFrame
        df = pd.DataFrame(sample_data, columns=headers)
        
        # Write to Excel
        df.to_excel(filename, index=False, sheet_name="Credentials")
        print(f"✓ Created new template: {filename}")
        
        # Format the Excel file
        format_excel_file(filename)
        
    except Exception as e:
        print(f"✗ Error creating template: {e}")


def add_headers_to_existing_file(filename="System_IPs.xlsx"):
    """Add headers to an existing Excel file if they don't exist"""
    try:
        if not os.path.exists(filename):
            print(f"ℹ File '{filename}' not found. Creating new template...\n")
            create_excel_template(filename)
            return
        
        # Read the existing file
        df = pd.read_excel(filename)
        
        # Check if headers already exist
        expected_headers = ["ILO_IP", "ILO_USERNAME", "ILO_PASSWORD"]
        
        if list(df.columns) == expected_headers:
            print(f"✓ File '{filename}' already has correct headers!")
            print(f"  Headers: {', '.join(expected_headers)}")
            print(f"  Total records: {len(df)}")
            return
        
        # If no headers or wrong headers, recreate with headers
        print(f"ℹ File '{filename}' found but headers are missing or incorrect.")
        
        # Read data without headers
        df_raw = pd.read_excel(filename, header=None)
        
        # Create new DataFrame with proper headers
        df_new = pd.DataFrame(df_raw.values, columns=expected_headers)
        
        # Write back to Excel
        df_new.to_excel(filename, index=False, sheet_name="Credentials")
        print(f"✓ Added headers to '{filename}'")
        print(f"  Headers: {', '.join(expected_headers)}")
        print(f"  Total records: {len(df_new)}")
        
        # Format the Excel file
        format_excel_file(filename)
        
    except Exception as e:
        print(f"✗ Error processing file: {e}")


def format_excel_file(filename="System_IPs.xlsx"):
    """Apply professional formatting to the Excel file"""
    try:
        wb = load_workbook(filename)
        ws = wb.active
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        thin_border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000")
        )
        
        # Format header row
        for col_num, header in enumerate(ws[1], 1):
            header.font = header_font
            header.fill = header_fill
            header.alignment = header_alignment
            header.border = thin_border
        
        # Format data rows
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
            for cell in row:
                cell.alignment = data_alignment
                cell.border = thin_border
        
        # Set column widths
        ws.column_dimensions['A'].width = 18  # ILO_IP
        ws.column_dimensions['B'].width = 18  # ILO_USERNAME
        ws.column_dimensions['C'].width = 18  # ILO_PASSWORD
        
        # Set row height for header
        ws.row_dimensions[1].height = 25
        
        wb.save(filename)
        print(f"✓ Applied professional formatting to '{filename}'")
        
    except Exception as e:
        print(f"✗ Error formatting Excel file: {e}")


def main():
    """Main function"""
    print("=" * 70)
    print("SYSTEM_IPS.XLSX - HEADER SETUP SCRIPT")
    print("=" * 70)
    print()
    
    filename = "System_IPs.xlsx"
    
    if os.path.exists(filename):
        print(f"Found existing file: {filename}\n")
        add_headers_to_existing_file(filename)
    else:
        print(f"File '{filename}' not found.\n")
        create_excel_template(filename)
    
    print()
    print("=" * 70)
    print("NEXT STEPS:")
    print("=" * 70)
    print("1. Open 'System_IPs.xlsx' in Excel or LibreOffice Calc")
    print("2. Replace sample data with your actual iLO/iDRAC credentials:")
    print("   - ILO_IP: Your server's iLO/iDRAC IP address")
    print("   - ILO_USERNAME: Username for iLO/iDRAC access")
    print("   - ILO_PASSWORD: Password for iLO/iDRAC access")
    print("3. Save the file")
    print("4. Run: python result.py")
    print("=" * 70)


if __name__ == "__main__":
    main()
